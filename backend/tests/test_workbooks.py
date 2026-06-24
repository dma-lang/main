"""FR-1 workbook parser: pillar-wise xlsx -> seed, with subcap-ID governance.

Pure unit tests over in-memory workbooks (no DB): the header variants that actually occur across
v5/v7, row skipping, exact-duplicate dedupe, and the ID-collision rules — reconcile by name against
the governing register, conflict when the register cannot place it, NEVER reuse/recycle/mint ids.
"""

from __future__ import annotations

import io
import zipfile
from typing import Any

import pytest

from app.services import workbooks


def _book(sheets: dict[str, list[list[Any]]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zip(files: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for n, b in files.items():
            zf.writestr(n, b)
    return buf.getvalue()


V7_HDR = [
    "Sub_Cap_ID",
    "Sub_Cap_Name",
    "Description",
    "Tier",
    "Category",
    "L1_Capability",
    "Solution_Type",
    "Zennify_Status",
]
V5_HDR = [
    "Sub-Cap ID",
    "Sub-Capability",
    "Description",
    "Tier",
    "Category ID",
    "Category Name",
    "Capability",
]
# the v5 Pillar-3 workbook names its columns differently again
V5_P3_HDR = [
    "Sub-Cap ID",
    "Sub-Capability Name",
    "Description",
    "Tier",
    "Capability ID",
    "Capability Name",
]


def test_parses_v5_and_v7_header_variants() -> None:
    data = _zip(
        {
            "Pillar 1 v7.xlsx": _book(
                {
                    "1_Overview": [["ignored"]],
                    "2_Capability_Map": [
                        V7_HDR,
                        [
                            "P1C1.1.1",
                            "Vision Setting",
                            "d",
                            "Core",
                            "P1C1 - Digital Strategy",
                            "Strategy",
                            "Advisory",
                            "Active",
                        ],
                    ],
                }
            ),
            "Pillar 2 v5.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        [
                            "P2C1.1",
                            "Journey Mapping",
                            "d",
                            "Core",
                            "P2C1",
                            "Experience Design",
                            "CX",
                        ],
                    ]
                }
            ),
            "Pillar 3 v5.xlsx": _book(
                {
                    "Capability Map": [
                        V5_P3_HDR,
                        ["P3C1.1", "Workflow Automation", "d", "Core", "P3C1", "Ops Automation"],
                    ]
                }
            ),
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest")
    ids = {s["id"] for s in out["subcaps"]}
    assert ids == {"P1C1.1.1", "P2C1.1", "P3C1.1"}
    assert set(out["pillars"]) == {"P1", "P2", "P3"}
    by_id = {s["id"]: s for s in out["subcaps"]}
    # v7: category name split out of "P1C1 - Digital Strategy"; cluster from L1_Capability
    assert by_id["P1C1.1.1"]["catId"] == "P1C1"
    assert by_id["P1C1.1.1"]["catName"] == "Digital Strategy"
    assert by_id["P1C1.1.1"]["cluster"] == "Strategy"
    # v5: explicit Category ID/Name columns; cluster from Capability
    assert by_id["P2C1.1"]["catName"] == "Experience Design"
    assert by_id["P2C1.1"]["cluster"] == "CX"
    # v5 P3 variant: no category columns -> cat id derives from the subcap id itself
    assert by_id["P3C1.1"]["catId"] == "P3C1"
    assert by_id["P3C1.1"]["cluster"] == "Ops Automation"
    assert out["skipped_rows"] == 0
    assert out["id_reconciliations"] == [] and out["id_conflicts"] == []
    # the DETECTED SCHEMA the onboarding review step shows: sheet + column->field per workbook
    detail = {d["file"]: d for d in out["workbooks_detail"]}
    assert detail["Pillar 1 v7.xlsx"]["sheet"] == "2_Capability_Map"
    cols = {c["source"]: c["field"] for c in detail["Pillar 1 v7.xlsx"]["columns"]}
    assert cols["Sub_Cap_ID"] == "id" and cols["Sub_Cap_Name"] == "name"
    assert detail["Pillar 1 v7.xlsx"]["subcaps_parsed"] == 1
    assert detail["Pillar 2 v5.xlsx"]["sheet"] == "Capability Map"
    assert {c["source"]: c["field"] for c in detail["Pillar 2 v5.xlsx"]["columns"]}[
        "Sub-Cap ID"
    ] == "id"


def test_skips_rows_without_a_subcap_id_or_name() -> None:
    data = _zip(
        {
            "Pillar 1.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        ["", "No id", "", "", "", "", ""],
                        ["NOT-AN-ID", "Bad id", "", "", "", "", ""],
                        ["P1C1.1", "", "", "", "", "", ""],  # no name
                        ["P1C1.2", "Kept", "", "", "P1C1", "Strategy", ""],
                    ]
                }
            )
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest")
    assert [s["id"] for s in out["subcaps"]] == ["P1C1.2"]
    assert out["skipped_rows"] == 3  # counted, never invented


def test_exact_duplicate_rows_dedupe() -> None:
    data = _zip(
        {
            "Pillar 2.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        ["P2C1.1", "Journey Mapping", "", "", "P2C1", "Experience", ""],
                        ["P2C1.1", "journey mapping", "", "", "P2C1", "Experience", ""],
                    ]
                }
            )
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest")
    assert len(out["subcaps"]) == 1
    assert out["duplicate_rows"] == 1
    assert out["id_conflicts"] == []  # same subcap, not a collision


def test_id_collision_reconciles_by_name_against_the_register() -> None:
    # Two DIFFERENT subcaps stamped with one id in the source (the real v5 P2C3.2.IC1 case).
    # The collider is reconciled to the governing version's id for that name — never re-minted.
    data = _zip(
        {
            "Pillar 2.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        ["P2C3.2.IC1", "Policy Self-Service", "", "", "P2C3", "Insurance", ""],
                        ["P2C3.2.IC1", "AI Claims Estimation", "", "", "P2C3", "Insurance", ""],
                    ]
                }
            )
        }
    )
    register = {"AI Claims Estimation": "P2C3.2.IC2"}
    out = workbooks.parse_catalogue_zip(data, "vtest", id_register=register)
    by_id = {s["id"]: s for s in out["subcaps"]}
    assert set(by_id) == {"P2C3.2.IC1", "P2C3.2.IC2"}  # both kept, ids never reused
    assert by_id["P2C3.2.IC2"]["name"] == "AI Claims Estimation"
    assert out["id_reconciliations"] == [
        {
            "source_id": "P2C3.2.IC1",
            "assigned_id": "P2C3.2.IC2",
            "name": "AI Claims Estimation",
            "via": "register",
        }
    ]
    assert out["id_conflicts"] == []


def test_id_collision_without_a_register_match_is_a_human_conflict() -> None:
    data = _zip(
        {
            "Pillar 2.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        ["P2C1.1", "First Owner", "", "", "P2C1", "Experience", ""],
                        ["P2C1.1", "Unplaceable Collider", "", "", "P2C1", "Experience", ""],
                    ]
                }
            )
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest")  # no register at all
    assert [s["id"] for s in out["subcaps"]] == ["P2C1.1"]
    assert out["subcaps"][0]["name"] == "First Owner"
    assert out["id_conflicts"] == [
        {"source_id": "P2C1.1", "name": "Unplaceable Collider", "file": "Pillar 2.xlsx"}
    ]


def test_register_id_already_owned_is_a_conflict_not_a_reuse() -> None:
    # The register points the collider at an id the upload already uses — assigning it would
    # recycle an id, so it must land as a conflict for a human instead.
    data = _zip(
        {
            "Pillar 2.xlsx": _book(
                {
                    "Capability Map": [
                        V5_HDR,
                        ["P2C1.1", "Alpha", "", "", "P2C1", "Experience", ""],
                        ["P2C1.2", "Beta", "", "", "P2C1", "Experience", ""],
                        ["P2C1.1", "Beta", "", "", "P2C1", "Experience", ""],
                    ]
                }
            )
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest", id_register={"Beta": "P2C1.2"})
    assert {s["id"] for s in out["subcaps"]} == {"P2C1.1", "P2C1.2"}
    assert out["id_reconciliations"] == []
    assert out["id_conflicts"] == [{"source_id": "P2C1.1", "name": "Beta", "file": "Pillar 2.xlsx"}]


def test_consolidated_and_non_pillar_files_are_ignored() -> None:
    pillar = _book({"Capability Map": [V5_HDR, ["P1C1.1", "Kept", "", "", "P1C1", "Strategy", ""]]})
    junk = _book({"Capability Map": [V5_HDR, ["P4C1.1", "Should not appear", "", "", "", "", ""]]})
    data = _zip(
        {
            "Pillar 1.xlsx": pillar,
            "Consolidated Pillar Map.xlsx": junk,
            "Archived Pillar 4.xlsx": junk,
            "Random Notes.xlsx": junk,
            "notes.txt": b"ignored",
        }
    )
    out = workbooks.parse_catalogue_zip(data, "vtest")
    assert [s["id"] for s in out["subcaps"]] == ["P1C1.1"]


def test_actionable_errors() -> None:
    with pytest.raises(ValueError, match="no pillar .xlsx workbooks"):
        workbooks.parse_catalogue_zip(_zip({"notes.txt": b"x"}), "vtest")
    # a member that is not a real workbook is a clean, named error — never a 500
    with pytest.raises(ValueError, match="Pillar 1.xlsx.*not a readable"):
        workbooks.parse_catalogue_zip(_zip({"Pillar 1.xlsx": b"not an xlsx"}), "vtest")
    no_sheet = _zip({"Pillar 1.xlsx": _book({"Wrong Tab": [["a"]]})})
    with pytest.raises(ValueError, match="no Capability Map sheet"):
        workbooks.parse_catalogue_zip(no_sheet, "vtest")
    bad_hdr = _zip({"Pillar 1.xlsx": _book({"Capability Map": [["Code", "Title"]]})})
    with pytest.raises(ValueError, match="headers not recognised"):
        workbooks.parse_catalogue_zip(bad_hdr, "vtest")


STORY_HDR = [
    "Story_Key",
    "Source_Type",
    "Sub_Cap_ID",
    "Sub_Cap_Name",
    "Story_Summary",
    "X_Acceptance_Criteria",
    "Y_Solution_Design",
    "Match_Confidence",
]


def test_synthetic_story_parse_excludes_jira_rows() -> None:
    data = _zip(
        {
            "Pillar 1.xlsx": _book(
                {
                    "Capability Map": [V5_HDR],
                    "3_User_Stories_Catalogue": [
                        STORY_HDR,
                        ["JIRA-1", "jira_completed", "P1C1.1", "n", "real row", "", "", "HIGH"],
                        [
                            "GEN-P1C1.1-01",
                            "gen_stories_v1",
                            "P1C1.1",
                            "n",
                            "made up",
                            "ac",
                            "sd",
                            "Medium",
                        ],
                        [
                            "PUB-P1C1.1-01",
                            "use_case_derived_public_validated",
                            "P1C1.1",
                            "n",
                            "derived",
                            "",
                            "",
                            "",
                        ],
                        ["", "gen_stories_v1", "P1C1.1", "n", "keyless -> skipped", "", "", ""],
                    ],
                }
            )
        }
    )
    rows = workbooks.parse_synthetic_stories_zip(data)
    assert [r["story_key"] for r in rows] == ["GEN-P1C1.1-01", "PUB-P1C1.1-01"]
    assert all(r["source_type"] != "jira_completed" for r in rows)
    assert rows[0]["confidence_level"] == "Medium"  # normalised to the enum at ingest, not here
    assert rows[0]["source_file"] == "Pillar 1.xlsx"


def test_story_xlsx_reads_the_actual_real_client_sheet() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    default = wb.active
    assert default is not None
    default.title = "Summary"
    default.append(["junk"])
    ws = wb.create_sheet(title="Actual (Real Client)")
    ws.append(["story_key", "summary", "sub_cap_id"])
    ws.append(["PROJ-1", "real jira story", "P1C1.1"])
    ws.append([None, "keyless row skipped", "P1C1.1"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    rows = workbooks.parse_story_xlsx(buf)
    assert len(rows) == 1
    assert rows[0]["story_key"] == "PROJ-1"
    assert rows[0]["is_synthetic"] is False  # by definition of the sheet
