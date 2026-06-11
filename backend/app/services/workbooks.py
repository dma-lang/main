"""Pillar-wise workbook parsing (FR-1) — the real xlsx -> catalogue pipeline.

Parses the uploaded pillar-wise capability maps (a ZIP of per-pillar .xlsx workbooks) into the
exact seed shape ``bring_version_online()`` consumes, and the v7 workbooks' embedded user-story
catalogue into SYNTHETIC story rows (everything not ``jira_completed`` — gen_stories_v1,
gen_synthesized_gap_fill, use_case_derived_public_validated). The REAL Jira corpus comes only
from the Full Story Catalog xlsx (sheet "Actual (Real Client)") — the two are never mixed, so
analysis can be Jira-only while synthetic rows stay visible and labelled.

Robustness over assumptions: sheet names and column headers differ across versions (v5
"Capability Map"/"Sub-Cap ID" vs v7 "2_Capability_Map"/"Sub_Cap_ID"), so both are resolved by
tolerant aliasing; rows without a subcap/story id are skipped (counted, never invented); pillar
ids derive from the subcap id itself; consolidated/archived workbooks in an upload are ignored.
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import Any, BinaryIO

PILLAR_NAMES = {
    "P1": "Strategy, Governance & Culture",
    "P2": "Customer Experience & Engagement",
    "P3": "Process Automation & Operations",
    "P4": "Data & AI Enablement",
}

_SUBCAP_ID_RE = re.compile(r"^P[1-4]C\d", re.IGNORECASE)


def _norm_header(h: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


# header aliases -> canonical field (normalised: lowercase, alnum only)
_CAP_ALIASES: dict[str, str] = {
    "subcapid": "id",
    "subcapability": "name",
    "subcapname": "name",
    "subcapabilityname": "name",  # v5 P3 variant
    "description": "desc",
    "tier": "tier",
    "categoryid": "catId",
    "categoryname": "catName",
    "category": "category_raw",  # v7: may carry "P1C1 - Name" or just the name
    "capid": "capId",
    "capabilityid": "capId",
    "capability": "cluster",
    "capabilityname": "cluster",
    "l1capability": "cluster",
    "solutiontype": "sol",
    "zennifystatus": "status",
    "primaryproducts": "platforms_raw",
    "l3platformsaddressingsubcap": "platforms_raw",
    # per-subcap REAL Jira story references the catalogue authors mapped (all four v7 pillar
    # workbooks carry this column) — they become story_subcap_carry links at carry-forward
    "storyrefswithuclinks": "story_refs",
    "storyrefs": "story_refs",
    "jirastoryrefs": "story_refs",
}

# Jira issue keys inside a refs cell ("BCFSC-1273, EBT-277 | MCUDA-13 …")
_JIRA_KEY_RE = re.compile(r"\b[A-Z][A-Z0-9]{1,9}-\d{1,5}\b")

# The 9 canonical subverticals as the v7 VC-mapping sheet names them -> data codes (these match
# the corpus story_sv_code / tier suffixes: RB, CU, CL, CIB, FC=Farm Credit, AM, RIA, IC, IB).
_VC_SV_CODES: dict[str, str] = {
    "retailbanking": "RB",
    "creditunions": "CU",
    "commerciallending": "CL",
    "corpinvestmentbanking": "CIB",
    "farmcreditaglending": "FC",
    "assetwealthmanagement": "AM",
    "riabrokerdealer": "RIA",
    "insurancecarriers": "IC",
    "insurancebrokerages": "IB",
}


def parse_vc_mapping_zip(data: bytes) -> dict[str, Any]:
    """The per-subcap × per-subvertical VALUE-CHAIN mapping the v7 pillar workbooks ship
    (sheet ``21_VC_Mapping_PerSubcap``): each subvertical column holds the named value-chain
    stage(s) the subcap belongs to ("▌ BACK OFFICE OPS", …). Returns
    ``{svs: {code: display}, mapping: [{subcap_id, sv, stages: [...]}], stage_order: {sv: [...]}}``
    — stage order preserved as first-seen per subvertical (the sheet lists subcaps in chain
    order), names kept verbatim (the user-facing labels), deterministic output. Workbooks without
    the sheet yield an empty mapping (v5 inherits the reference's at provision)."""
    books = _pillar_workbooks(data)
    mapping: dict[tuple[str, str], list[str]] = {}
    svs: dict[str, str] = {}
    stage_order: dict[str, list[str]] = {}
    for fname, blob in books:
        wb = _load_book(fname, blob)
        ws = _sheet(wb, "vcmappingpersubcap", "vcmapping")
        if ws is None:
            continue
        rows = ws.iter_rows(values_only=True)
        header: list[str] | None = None
        for r in rows:  # the sheet has a banner block; the real header starts at 'Category'
            cells = [str(x or "") for x in r]
            if cells and cells[0] == "Category" and any("Sub_Cap_ID" in c for c in cells):
                header = cells
                break
        if header is None:
            continue
        sid_i = header.index("Sub_Cap_ID")
        sv_cols: list[tuple[int, str]] = []
        for i, h in enumerate(header):
            code = _VC_SV_CODES.get(_norm_header(h))
            if code:
                sv_cols.append((i, code))
                svs.setdefault(code, h.strip())
        for r in rows:
            if len(r) <= sid_i:
                continue
            sid = str(r[sid_i] or "").strip()
            if not _SUBCAP_ID_RE.match(sid):
                continue
            for i, code in sv_cols:
                cell = str(r[i] or "") if i < len(r) else ""
                # stages are "▌"-prefixed blocks; commas INSIDE a stage name are part of it.
                # "- (N/A)" / "Not applicable — …" cells mean the subcap has no stage in this
                # subvertical's chain — honest absence, not a stage.
                stages = [s.strip(" ▌\n\t") for s in cell.split("▌")]
                # "(applicable via X pattern)" is provenance, not part of the stage name —
                # stripping it collapses the cross-pattern variants into the one real stage
                stages = [
                    re.sub(r"\s*\(applicable via [^)]*\)\s*$", "", re.sub(r"\s+", " ", s)).strip()
                    for s in stages
                ]
                stages = [
                    s
                    for s in stages
                    if s
                    and not re.match(r"^-?\s*\(?n/?a\)?\.?$", s, re.I)
                    and not s.lower().startswith("not applicable")
                    and s.lower() not in ("none", "-")
                ]
                if not stages:
                    continue
                mapping.setdefault((sid, code), [])
                order = stage_order.setdefault(code, [])
                for st in stages:
                    if st not in mapping[(sid, code)]:
                        mapping[(sid, code)].append(st)
                    if st not in order:
                        order.append(st)
    return {
        "svs": svs,
        "mapping": [
            {"subcap_id": sid, "sv": sv, "stages": stages}
            for (sid, sv), stages in sorted(mapping.items())
        ],
        "stage_order": stage_order,
    }


_STORY_ALIASES: dict[str, str] = {
    "storykey": "story_key",
    "sourcetype": "source_type",
    "subcapid": "sub_cap_id",
    "subcapname": "sub_cap_name",
    "storysummary": "summary",
    "xacceptancecriteria": "ac_text",
    "ysolutiondesign": "solution_design_text",
    "matchconfidence": "confidence_level",
}


def _sheet(wb: Any, *want: str) -> Any | None:
    """First sheet whose normalised name contains any wanted token."""
    for name in wb.sheetnames:
        n = _norm_header(name)
        if any(w in n for w in want):
            return wb[name]
    return None


def _index(header: tuple[Any, ...], aliases: dict[str, str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        key = aliases.get(_norm_header(h))
        if key and key not in out:
            out[key] = i
    return out


def _cell(row: tuple[Any, ...], idx: dict[str, int], key: str) -> str:
    i = idx.get(key)
    if i is None or i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


def _pillar_workbooks(data: bytes) -> list[tuple[str, bytes]]:
    """The per-pillar workbooks inside an upload zip (consolidated/archived files ignored)."""
    zf = zipfile.ZipFile(io.BytesIO(data))
    out: list[tuple[str, bytes]] = []
    for info in zf.infolist():
        base = info.filename.rsplit("/", 1)[-1]
        low = base.lower()
        if not low.endswith(".xlsx") or base.startswith((".", "~")):
            continue
        if "consolidated" in low or "archived" in low:
            continue
        if "pillar" in low:
            out.append((base, zf.read(info)))
    return sorted(out)


def _load_book(fname: str, blob: bytes) -> Any:
    """Open one workbook, translating openpyxl's internal errors (an .xlsx is itself a zip) into
    the parser's actionable ValueError so a corrupt member is a clean 400, never a 500."""
    import openpyxl

    try:
        return openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as exc:  # zipfile.BadZipFile, InvalidFileException, truncated streams …
        raise ValueError(f"{fname}: not a readable .xlsx workbook ({exc})") from exc


def parse_catalogue_zip(
    data: bytes, version: str, id_register: dict[str, str] | None = None
) -> dict[str, Any]:
    """ZIP of pillar workbooks -> the provisioning seed shape
    {pillars, catNames, subcaps:[...]} + governance fields.

    ID GOVERNANCE (subcap ids are never reused, recycled, or minted here): when two DIFFERENT
    subcaps collide on one id in the source, the first occurrence keeps it; the collider is
    RECONCILED by name against ``id_register`` (the governing version's name -> id map, e.g. v7)
    and carries that authoritative id — recorded in ``id_reconciliations`` and written to the
    version crosswalk at provision. A collider the register cannot place lands in
    ``id_conflicts`` for a HUMAN to reconcile (kept out of the seed, never silently dropped,
    never given an invented id). Exact repeats (same id, same name) dedupe as ``duplicate_rows``.
    Raises ValueError with an actionable message when nothing parseable is found."""
    books = _pillar_workbooks(data)
    if not books:
        raise ValueError("no pillar .xlsx workbooks found in the zip")
    subcaps: list[dict[str, Any]] = []
    cat_names: dict[str, str] = {}
    seen: dict[str, str] = {}  # id -> name of the row that owns it
    register = {k.strip().lower(): v for k, v in (id_register or {}).items()}
    reconciliations: list[dict[str, str]] = []
    conflicts: list[dict[str, str]] = []
    detail: list[dict[str, Any]] = []  # per-workbook detected schema, for the automap review step
    duplicates = 0
    skipped = 0
    for fname, blob in books:
        wb = _load_book(fname, blob)
        ws = _sheet(wb, "capabilitymap")
        if ws is None:
            raise ValueError(f"{fname}: no Capability Map sheet (sheets: {wb.sheetnames[:6]}…)")
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"{fname}: Capability Map sheet is empty")
        idx = _index(header, _CAP_ALIASES)
        if "id" not in idx or "name" not in idx:
            raise ValueError(
                f"{fname}: Capability Map headers not recognised "
                f"(need a Sub-Cap ID and Sub-Capability/Sub_Cap_Name column)"
            )
        # DETECTED SCHEMA for the human review step: which sheet matched and how each source
        # column maps to a canonical field — plus the headers the automap did not recognise
        # (shown, never silently dropped).
        columns: list[dict[str, Any]] = [
            {"source": str(h).strip(), "field": _CAP_ALIASES[_norm_header(h)]}
            for h in header
            if h is not None and str(h).strip() and _norm_header(h) in _CAP_ALIASES
        ]
        unmapped = [
            str(h).strip()
            for h in header
            if h is not None and str(h).strip() and _norm_header(h) not in _CAP_ALIASES
        ]
        book_detail: dict[str, Any] = {
            "file": fname,
            "sheet": ws.title,
            "columns": columns,
            "unmapped_headers": unmapped,
            "subcaps_parsed": 0,
            "other_sheets": [n for n in wb.sheetnames if n != ws.title],
        }
        detail.append(book_detail)
        # Per-column automap evidence (the prototype's review contract): fill rate, format
        # validity, sample values — combined with the exact-alias header match into a confidence.
        stats: dict[str, dict[str, Any]] = {
            c["field"]: {"filled": 0, "ok": 0, "samples": []} for c in columns
        }
        rows_scanned = 0
        for row in rows:
            sid = _cell(row, idx, "id")
            name = _cell(row, idx, "name")
            rows_scanned += 1
            for c in columns:
                val = _cell(row, idx, c["field"])
                if val:
                    st = stats[c["field"]]
                    st["filled"] += 1
                    if len(st["samples"]) < 3:
                        st["samples"].append(val[:60])
                    if c["field"] != "id" or _SUBCAP_ID_RE.match(val):
                        st["ok"] += 1
            if not _SUBCAP_ID_RE.match(sid) or not name:
                skipped += 1
                continue
            if sid in seen:
                if seen[sid].strip().lower() == name.strip().lower():
                    duplicates += 1  # exact repeat of the same subcap — safe to drop one
                    continue
                # ID COLLISION: a different subcap under an already-owned id. Reconcile by name
                # against the governing register; ids are never reused, recycled, or minted here.
                gov = register.get(name.strip().lower())
                if gov and gov not in seen:
                    reconciliations.append(
                        {"source_id": sid, "assigned_id": gov, "name": name, "via": "register"}
                    )
                    sid = gov
                else:
                    conflicts.append({"source_id": sid, "name": name, "file": fname})
                    continue
            seen[sid] = name
            cat_id = _cell(row, idx, "catId") or sid.split(".", 1)[0]
            cat_name = _cell(row, idx, "catName")
            if not cat_name:
                raw = _cell(row, idx, "category_raw")
                # v7 style "P1C1 - Digital Strategy" or a bare name
                cat_name = raw.split(" - ", 1)[-1].strip() if raw else cat_id
            cat_names.setdefault(cat_id, cat_name)
            cluster = _cell(row, idx, "cluster") or cat_name
            subcaps.append(
                {
                    "id": sid,
                    "name": name,
                    "catId": cat_id,
                    "catName": cat_names[cat_id],
                    "cluster": cluster,
                    "tier": _cell(row, idx, "tier") or None,
                    "desc": _cell(row, idx, "desc") or None,
                    "sol": _cell(row, idx, "sol") or None,
                    "status": _cell(row, idx, "status") or None,
                    "life": "stable",
                    "comp": 0,
                    # the catalogue's own Jira story references for this subcap (deduped, sorted
                    # for a deterministic seed) — linked against the corpus at carry-forward
                    "story_refs": sorted(set(_JIRA_KEY_RE.findall(_cell(row, idx, "story_refs")))),
                }
            )
            book_detail["subcaps_parsed"] += 1
        # finalise the per-column evidence: header match (exact alias) is a strong prior; fill
        # rate and format validity are measured over THIS workbook's rows; 3 samples shown.
        for c in columns:
            st = stats[c["field"]]
            fill = (st["filled"] / rows_scanned) if rows_scanned else 0.0
            fmt = (st["ok"] / st["filled"]) if st["filled"] else 0.0
            c["samples"] = st["samples"]
            c["signals"] = {
                "header_match": 1.0,  # mapped via an exact, curated alias — never fuzzy-guessed
                "fill_rate": round(fill, 2),
                "format_valid": round(fmt, 2),
                "rows_scanned": rows_scanned,
            }
            c["confidence"] = round(0.5 + 0.3 * fill + 0.2 * fmt, 2)
    if not subcaps:
        raise ValueError("the pillar workbooks contained no subcap rows")
    pillars = {
        pid: {"name": PILLAR_NAMES.get(pid, pid)} for pid in sorted({s["id"][:2] for s in subcaps})
    }
    return {
        "version": version,
        "pillars": pillars,
        "catNames": cat_names,
        "subcaps": subcaps,
        "skipped_rows": skipped,
        "duplicate_rows": duplicates,
        "id_reconciliations": reconciliations,
        "id_conflicts": conflicts,
        "workbooks_detail": detail,  # per-book detected schema, for the automap review step
        "relations_detected": _detect_relations(detail),
    }


def _detect_relations(detail: list[dict[str, Any]]) -> list[dict[str, str]]:
    """The RELATIONSHIPS the backend schema needs, detected from what the workbooks actually
    contain (columns found + companion sheets present) — the onboarding relations review. These
    become real FKs / link tables at provision (relation_def); nothing here is guessed."""
    fields = {c["field"] for d in detail for c in d["columns"]}
    sheets = " ".join(_norm_header(n) for d in detail for n in d.get("other_sheets", []))
    rels: list[dict[str, str]] = [
        {
            "from": "subcap",
            "verb": "belongs_to",
            "to": "capability",
            "via": "Sub-Cap ID grammar (P<pillar>C<category>.<capability>…)",
        },
        {"from": "capability", "verb": "belongs_to", "to": "category", "via": "id grammar"},
        {"from": "category", "verb": "belongs_to", "to": "pillar", "via": "id grammar"},
    ]
    if "platforms_raw" in fields:
        rels.append(
            {
                "from": "subcap",
                "verb": "uses_platform",
                "to": "l3_platform",
                "via": "L3 platforms column",
            }
        )
    if "persona" in sheets:
        rels.append(
            {"from": "subcap", "verb": "has_persona", "to": "persona", "via": "Personas sheet"}
        )
    if "usecase" in sheets or "usecases" in sheets:
        rels.append(
            {"from": "subcap", "verb": "has_usecase", "to": "use_case", "via": "Use Cases sheet"}
        )
    if "maturity" in sheets:
        rels.append(
            {
                "from": "subcap",
                "verb": "custom",
                "to": "maturity_descriptor",
                "via": "Maturity sheet",
            }
        )
    if "stories" in sheets or "userstories" in sheets:
        rels.append(
            {
                "from": "story",
                "verb": "addresses",
                "to": "subcap",
                "via": "User Stories sheet (synthetic labelled; Jira corpus separate)",
            }
        )
    if "offering" in sheets or "offerings" in sheets:
        rels.append(
            {
                "from": "offering",
                "verb": "maps_to_offering",
                "to": "subcap",
                "via": "Offerings sheet",
            }
        )
    return rels


def parse_synthetic_stories_zip(data: bytes) -> list[dict[str, Any]]:
    """The v7 workbooks' embedded user-story catalogue, SYNTHETIC rows only — every Source_Type
    that is not ``jira_completed`` (the real corpus lives solely in the Full Story Catalog xlsx,
    so workbook copies of Jira rows are skipped rather than double-ingested)."""
    out: list[dict[str, Any]] = []
    for fname, blob in _pillar_workbooks(data):
        wb = _load_book(fname, blob)
        ws = _sheet(wb, "userstoriescatalogue", "storiescatalogue")
        if ws is None:
            continue  # a version without an embedded story tab (e.g. v5) simply has none
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            continue
        idx = _index(header, _STORY_ALIASES)
        if "story_key" not in idx:
            continue
        for row in rows:
            key = _cell(row, idx, "story_key")
            stype = _cell(row, idx, "source_type").lower()
            if not key or stype == "jira_completed":
                continue
            sid = _cell(row, idx, "sub_cap_id")
            out.append(
                {
                    "story_key": key,
                    "source_type": stype or "synthetic",
                    "sub_cap_id": sid or None,
                    "sub_cap_name": _cell(row, idx, "sub_cap_name") or None,
                    "summary": _cell(row, idx, "summary") or None,
                    "ac_text": _cell(row, idx, "ac_text") or None,
                    "solution_design_text": _cell(row, idx, "solution_design_text") or None,
                    "confidence_level": (_cell(row, idx, "confidence_level") or None),
                    "source_file": fname,
                }
            )
    return out


def parse_story_xlsx(fh: BinaryIO | str) -> list[dict[str, Any]]:
    """The Full Story Catalog xlsx (sheet "Actual (Real Client)") -> the REAL Jira corpus rows,
    columns passed through by their canonical snake_case names. is_synthetic is False by
    definition of the sheet; rows without a story_key are skipped."""
    import openpyxl

    wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "actual" in name.lower() or "real" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows)]
    out: list[dict[str, Any]] = []
    for row in rows:
        d = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        if not d.get("story_key"):
            continue
        d["is_synthetic"] = False
        out.append(d)
    return out
