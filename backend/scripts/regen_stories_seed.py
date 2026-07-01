"""Regenerate seed/stories.json.gz with the RICH Jira fields (R8).

The committed corpus seed was a lossy transform of the canonical 14,406-row Jira Full Story Catalog:
it kept only `summary` + the numeric sub-scores and DISCARDED the raw `description`, acceptance
criteria (`ac_text`), and `solution_design_text`, and carried no resolved client name. This script
re-derives the seed from the authoritative source workbook
(`Zennify_Full_Story_Catalog_with_Client_RESOLVED.xlsx`) by AUGMENTING the existing committed seed:
every existing short key stays byte-identical (so the carry-forward invariants — the story_key set,
the sub_cap_id mapping, and every score — are provably unchanged) and each row merely GAINS the rich
narrative + client fields under new short keys:

    desc  <- description            act <- ac_text            sdt <- solution_design_text (cap 8k)
    cn    <- client_name            said <- salesforce_account_id
    cmm   <- client_match_method    cmc  <- client_match_confidence

The workbook is verified to be an exact superset of the committed seed (same 14,406 story_keys, zero
mismatch on sub_cap_id / composite / ac_score); the script re-checks that coverage and aborts if it
ever drifts, so a future workbook can never silently perturb the corpus. Rich texts are capped at
``_TEXT_CAP`` chars (covers >99% in full) to bound the seed; empty fields are dropped; row order is
preserved for a clean, purely-additive git diff.

Run (engineer, once, not app runtime):
    uv run python scripts/regen_stories_seed.py <path-to.xlsx>
"""

from __future__ import annotations

import gzip
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_BACKEND = Path(__file__).resolve().parents[1]
_SEED = _BACKEND / "seed" / "stories.json.gz"
_TEXT_CAP = 8000  # per-field char cap for the rich blobs (bounds the committed seed size)

# xlsx column -> new short seed key (the RICH additions R8 layers on)
_RICH = {
    "description": "desc",
    "ac_text": "act",
    "solution_design_text": "sdt",
    "client_name": "cn",
    "salesforce_account_id": "said",
    "client_match_method": "cmm",
    "client_match_confidence": "cmc",
}
# the invariant fields we re-verify against the committed seed (must not drift)
_INVARIANT = {"sub_cap_id": "sc", "composite_score": "cs", "ac_score": "ac", "sd_score": "sd"}


def _cell(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _num(seed_val: Any, xl_val: Any) -> bool:
    try:
        return abs(float(seed_val) - float(xl_val)) < 1e-6
    except (TypeError, ValueError):
        return (seed_val is None and xl_val is None) or str(seed_val) == str(xl_val)


def main(xlsx_path: str) -> int:
    seed: list[dict[str, Any]] = json.loads(gzip.open(_SEED, "rt", encoding="utf-8").read())
    by_key = {row["k"]: row for row in seed}

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    for col in [*_RICH, *_INVARIANT, "story_key"]:
        if col not in idx:
            raise SystemExit(f"workbook is missing required column {col!r}")

    seen: set[str] = set()
    drift: list[str] = []
    added = 0
    for r in rows:
        key = _cell(r[idx["story_key"]])
        if not key:
            continue
        seen.add(key)
        target = by_key.get(key)
        if target is None:
            drift.append(f"{key} (in workbook, not in seed)")
            continue
        # guard: the invariant fields must match the committed seed exactly
        for col, sk in _INVARIANT.items():
            if not _num(target.get(sk), r[idx[col]]):
                drift.append(f"{key}.{sk}: seed={target.get(sk)!r} workbook={r[idx[col]]!r}")
        # augment with the rich + client fields (capped, empties dropped)
        for col, sk in _RICH.items():
            val = _cell(r[idx[col]])
            if val is not None:
                target[sk] = val[:_TEXT_CAP]
                added += 1

    missing = set(by_key) - seen
    if missing:
        drift.append(f"{len(missing)} committed-seed stories absent from the workbook")
    if drift:
        for d in drift[:15]:
            print("DRIFT:", d, file=sys.stderr)
        raise SystemExit(f"workbook drifted from the seed ({len(drift)} issues); aborting")

    payload = json.dumps(seed, ensure_ascii=False, separators=(",", ":"))
    with gzip.open(_SEED, "wt", encoding="utf-8") as fh:
        fh.write(payload)

    clients = {row["cn"] for row in seed if row.get("cn")}
    svs = {row.get("svn") for row in seed if row.get("svn")}
    subcaps = {row["sc"] for row in seed}
    with_desc = sum(1 for row in seed if row.get("desc"))
    with_ac = sum(1 for row in seed if row.get("act"))
    with_sd = sum(1 for row in seed if row.get("sdt"))
    print(f"rows={len(seed)}  rich fields added={added}")
    print(f"clients={len(clients)}  subverticals={len(svs)}  subcaps={len(subcaps)}")
    print(f"with description={with_desc}  ac_text={with_ac}  solution_design={with_sd}")
    print(f"seed size={_SEED.stat().st_size / 1e6:.2f} MB gz")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("usage: regen_stories_seed.py <path-to-story-catalog.xlsx>")
    raise SystemExit(main(sys.argv[1]))
